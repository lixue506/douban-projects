# -*- coding: utf-8 -*-
# 爬取用户所看电影信息
import re
import xlrd
import xlwt
import time
import requests
import openpyxl
from bs4 import BeautifulSoup

#save_information
def first_save_information():
    # 创建Excel文件,同时默认创建一个sheet
    f = openpyxl.Workbook()
    f.save('movie.xlsx')#激活创建的文件
    f = openpyxl.Workbook()
    # 创建表
    sheet_names = f.sheetnames
    sheet1 = f[sheet_names[0]]
    #写入一行数据
    row0 = ["电影名称", "评分","五星比例","四星比例","三星比例","二星比例","一星比例","类型","导演","演员一" ,"演员二","演员三"]
    for i in range(len(row0)):
        sheet1.cell(row=1,column=i+1,value=row0[i])
    f.save('movie.xlsx')#激活创建的文件

def get_url(url):
    try:
        d = {
            'Cookie':'bid=JoOO5Fbfy_U; douban-profile-remind=1; douban-fav-remind=1; ct=y; ps=y;\
             _ga=GA1.2.819158214.1555853957; loc-last-index-location-id="108296"; ll="108296"; \
             dbcl2="193741190:MGOdDJNKEdo"; ck=cz6U; frodotk="b11655a589bf862c134300f09d3e9284"; ap_v=0,6.0;\
              _vwo_uuid_v2=DBBFD1C4D4352E321CF39AA05ECDBC25C|ec2df2e7c16cc03c0576e19027669a33; __utmt=1;\
               push_noty_num=0; push_doumail_num=0;__utma=30149280.819158214.1555853957.1556616814.1556625559.20;\
                __utmb=30149280.4.10.1556625559; __utmc=30149280; __utmz=30149280.1556616814.19.6.utmcsr=douban.\
                com|utmccn=(referral)|utmcmd=referral|utmcct=/; __utmv=30149280.19374',
            'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36 SE 2.X MetaSr 1.0'}
        r = requests.get(url=url, headers=d)
        if r.status_code == 200:
            return r.text
    except:
        print("爬取失败！")
        return None

def save_information(name,rate,stars,tag,dictor,actors):
    #打开已有表
    f = openpyxl.load_workbook("movie.xlsx")
    #定位sheet页,打开第一个工作表
    sheet_names = f.sheetnames
    worksheet1 = f[sheet_names[0]]
    #写入文件
    col = worksheet1.max_row + 1
    row = 1
    worksheet1.cell(row=col,column=row,value=name)
    row += 1
    worksheet1.cell(row=col,column=row,value=rate)
    row += 1
    for star in stars:
        worksheet1.cell(row=col,column=row,value=star.text)
        if row == 7:
            row += 1
            break
        else:
            row += 1
    worksheet1.cell(row=col,column=row,value=tag)
    row += 1
    worksheet1.cell(row=col,column=row,value=dictor)
    row += 1
    #仅写入三个演员名称，写到三个或者不足三个跳出
    for actor in actors:
        worksheet1.cell(row=col,column=row,value=actor.text)
        if row == 12:
            break
        else:
            row += 1
    f.save('movie.xlsx')

#爬取电影信息
def parser_html(html):
    soup = BeautifulSoup(html, "html.parser")
    #name
    name = soup.find('span',attrs={'property':"v:itemreviewed"}).text
    #rate
    rate = soup.find('strong',attrs={'class':"ll rating_num",'property':"v:average"}).text
    # 五星评分
    stars = soup.find_all('span', attrs={'class': "rating_per"})
    # tag只写入一个标签
    tag = soup.find('span', attrs={'property': "v:genre"} ).text
    #dictor
    dictor = soup.find('a',attrs={'rel':"v:directedBy"}).text
    # actor
    actors = soup.find_all('a',attrs={'rel':"v:starring"})
    save_information(name,rate,stars,tag,dictor,actors)
    time.sleep(1)

def file_read_url():
    with open('url_old.txt','r') as f:
        while True:
            line = f.readline().strip('\n')
            if line == '':
                break
            else:
                html = get_url(line)
                parser_html(html)
                print("已写入！")

if __name__ == "__main__":
    first_save_information()
    file_read_url()

